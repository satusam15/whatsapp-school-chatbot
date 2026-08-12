"""
manage_students.py

Simple command-line tool to add/list students once the school hands over
real data. No code changes needed to onboard a new student — just run this.

Usage:
    python3 manage_students.py add
        (interactive prompts for one student + parent number)

    python3 manage_students.py list
        (prints every student + linked parent number currently in the DB)

    python3 manage_students.py sync-sheet master_sheet.csv
        (THE MAIN COMMAND — handles the school's real export format. Use
        this every time the school sends an updated master sheet, for
        first import AND every exam cycle after. See SYNC_SHEET_FORMAT below.)

    python3 manage_students.py import students.csv
        (older/simpler format, still works if you're not using the
        school's real export — see CSV_FORMAT below.)

    python3 manage_students.py update-marks marks.csv
        (older format for marks-only uploads matched by roll_number —
        sync-sheet supersedes this for the school's actual workflow.)

SYNC_SHEET_FORMAT — matches the school's real export as-is:
    A few header/title rows are fine before the real header row. The real
    header row must contain a column with "Student Name" and one with
    "Mobile No" (case-insensitive, minor spacing differences OK). Expected
    columns: Student Name, Mobile No, Standard, Attendence, and ONE marks
    column whose header contains "MARKS" (e.g. "FA 1 MARKS", "FA 2 MARKS")
    — that header is what tells the system which exam this is. A "%"
    column and a "GRADE" column are used if present.

    Each row is matched to an existing student by (name + mobile number)
    together. If found: attendance/grade are refreshed and marks are
    added/updated for the detected exam. If not found: a new student is
    created with an auto-assigned school ID (MKS0001 style).

    Re-running with the same file, or a corrected version of it, is safe —
    nothing gets duplicated.

CSV_FORMAT for `import` (header row required, exact column names):
    parent_phone,student_name,grade,roll_number,attendance,exam_schedule,fee_status

    Example row:
    918861388136,Savita,Grade 8,42,94%,"Maths: Mon 9AM",Paid

MARKS_CSV_FORMAT for `update-marks` (header row required, exact column names):
    roll_number,exam_name,marks,percentage,grade

    Example row:
    42,FA2,115,92,A+
"""

import sys
import re
import csv
import openpyxl
from db import (
    get_connection, init_db, upsert_marks,
    find_student_by_name_and_phone, update_student_fields, upsert_marks_by_student_id,
    delete_student
)


def add_student_interactive():
    print("Enter student details (leave blank to skip a field):")
    student_name = input("Student name: ").strip()
    grade = input("Grade: ").strip()
    roll_number = input("Roll number (recommended, used for later mark updates): ").strip()
    attendance = input("Attendance (e.g. 94%): ").strip()
    exam_schedule = input("Exam schedule: ").strip()
    fee_status = input("Fee status: ").strip()
    parent_phone = input("Parent WhatsApp number (country code, no +, no spaces, e.g. 917411477367): ").strip()

    if not student_name or not parent_phone:
        print("Student name and parent phone are required. Aborting.")
        return

    add_student(parent_phone, student_name, grade, attendance, exam_schedule, fee_status, roll_number)
    print(f"Added: {student_name} linked to {parent_phone}")


def add_student(parent_phone, student_name, grade, attendance, exam_schedule, fee_status, roll_number=None):
    from db import add_student_row
    add_student_row(parent_phone, student_name, grade, attendance, exam_schedule, fee_status, roll_number)


def list_students():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.id, s.student_name, s.grade, p.whatsapp_number
        FROM students s
        LEFT JOIN parents p ON p.student_id = s.id
        ORDER BY s.id
    """)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        print("No students in the database yet.")
        return

    print(f"{'ID':<4} {'Name':<20} {'Grade':<10} {'Parent Number'}")
    for row in rows:
        print(f"{row['id']:<4} {row['student_name']:<20} {row['grade'] or '':<10} {row['whatsapp_number'] or '(none)'}")


def _next_auto_roll_numbers(conn, count):
    """Generates `count` new sequential IDs like MKS0001, MKS0002... continuing
    from whatever's already been assigned, so repeated imports never collide."""
    cur = conn.cursor()
    cur.execute("SELECT roll_number FROM students WHERE roll_number LIKE 'MKS%'")
    existing = [row["roll_number"] for row in cur.fetchall()]
    numbers = []
    for r in existing:
        try:
            numbers.append(int(r[3:]))
        except (ValueError, IndexError):
            continue
    start = (max(numbers) + 1) if numbers else 1
    return [f"MKS{n:04d}" for n in range(start, start + count)]


def _find_header_row(rows):
    """Scans raw CSV rows for the real header row — the school's export has
    a few title/blank rows before it. Returns (header_index, header_list)."""
    for i, row in enumerate(rows):
        joined = " ".join(cell.strip().lower() for cell in row)
        if "student name" in joined and "mobile" in joined:
            return i, row
    return None, None


def _normalize_exam_name(marks_header: str) -> str:
    """'FA 1 MARKS' -> 'FA1', 'SA 2 Marks' -> 'SA2', etc."""
    cleaned = re.sub(r"marks", "", marks_header, flags=re.IGNORECASE)
    cleaned = re.sub(r"[^a-zA-Z0-9]", "", cleaned)
    return cleaned.upper() or "EXAM"


def sync_master_sheet(path):
    with open(path, newline="", encoding="utf-8") as f:
        raw_rows = list(csv.reader(f))

    header_idx, header = _find_header_row(raw_rows)
    if header is None:
        print("Couldn't find a header row containing 'Student Name' and 'Mobile No'. Aborting.")
        return

    def col_index(*keywords):
        for i, col in enumerate(header):
            col_lower = col.strip().lower()
            if all(kw in col_lower for kw in keywords):
                return i
        return None

    idx_name = col_index("student name")
    idx_mobile = col_index("mobile")
    idx_standard = col_index("standard")
    idx_attendance = col_index("attend")
    idx_marks = col_index("marks")
    idx_percent = next((i for i, c in enumerate(header) if c.strip() == "%"), None)
    idx_grade = col_index("grade")

    if idx_name is None or idx_mobile is None:
        print("Missing required 'Student Name' or 'Mobile No' column. Aborting.")
        return

    exam_name = _normalize_exam_name(header[idx_marks]) if idx_marks is not None else None
    if exam_name:
        print(f"Detected exam column: '{header[idx_marks]}' -> storing as exam '{exam_name}'")
    else:
        print("No marks column detected — will only sync names/grades/attendance, no marks.")

    data_rows = [r for r in raw_rows[header_idx + 1:] if any(c.strip() for c in r)]
    total = len(data_rows)
    print(f"Processing {total} rows (using a single database connection for speed)...")

    updated = 0
    created = 0
    skipped = 0
    reference_rows = []

    # ONE connection, reused for the entire sync — avoids reconnecting to a
    # remote database per student, which is what made large syncs painfully
    # slow (each connection is a fresh network round-trip to Supabase).
    conn = get_connection()
    cur = conn.cursor()

    # Preload existing roll numbers once, so we're not re-querying for every new student.
    cur.execute("SELECT roll_number FROM students WHERE roll_number LIKE 'MKS%'")
    existing_numbers = []
    for row in cur.fetchall():
        try:
            existing_numbers.append(int(row["roll_number"][3:]))
        except (ValueError, IndexError, TypeError):
            continue
    next_auto_num = (max(existing_numbers) + 1) if existing_numbers else 1

    for i, row in enumerate(data_rows, start=1):
        def get(idx):
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        name = get(idx_name)
        mobile = get(idx_mobile)
        standard = get(idx_standard)
        attendance = get(idx_attendance)
        marks_val = get(idx_marks)
        percent_val = get(idx_percent)
        grade_val = get(idx_grade)

        if not name or not mobile:
            print(f"Skipping row with missing name/mobile: {row}")
            skipped += 1
            continue

        if not (mobile.isdigit() and len(mobile) >= 10):
            print(f"Skipping row with malformed mobile number for '{name}': '{mobile}'")
            skipped += 1
            continue

        grade_label = f"Standard {standard}" if standard else None

        cur.execute("""
            SELECT s.* FROM parents p JOIN students s ON p.student_id = s.id
            WHERE p.whatsapp_number = %s AND LOWER(s.student_name) = LOWER(%s)
        """, (mobile, name))
        existing = cur.fetchone()

        if existing:
            if grade_label is not None:
                cur.execute("UPDATE students SET grade = %s WHERE id = %s", (grade_label, existing["id"]))
            if attendance:
                cur.execute("UPDATE students SET attendance = %s WHERE id = %s", (attendance, existing["id"]))
            student_id = existing["id"]
            roll_number = existing["roll_number"]
            updated += 1
        else:
            roll_number = f"MKS{next_auto_num:04d}"
            next_auto_num += 1
            cur.execute(
                "INSERT INTO students (student_name, grade, attendance, roll_number) "
                "VALUES (%s, %s, %s, %s) RETURNING id",
                (name, grade_label, attendance or None, roll_number)
            )
            student_id = cur.fetchone()["id"]
            cur.execute(
                "INSERT INTO parents (whatsapp_number, student_id) VALUES (%s, %s)",
                (mobile, student_id)
            )
            created += 1

        if exam_name:
            cur.execute("""
                INSERT INTO marks (student_id, exam_name, marks, percentage, grade)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (student_id, exam_name)
                DO UPDATE SET marks = EXCLUDED.marks,
                               percentage = EXCLUDED.percentage,
                               grade = EXCLUDED.grade,
                               recorded_at = CURRENT_TIMESTAMP
            """, (student_id, exam_name, marks_val or None, percent_val or None, grade_val or None))

        reference_rows.append((roll_number, name, mobile, grade_label or ""))

        if i % 50 == 0 or i == total:
            print(f"  ...{i}/{total} processed")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nSync complete: {created} new students added, {updated} existing students updated, {skipped} rows skipped.")

    ref_path = "assigned_ids_reference.csv"
    with open(ref_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["roll_number", "student_name", "parent_phone", "grade"])
        writer.writerows(reference_rows)
    print(f"Updated ID reference sheet: {ref_path}")


def _fill_merged_row(ws, row_num, max_col):
    """Reads a header row, filling in values for cells that are part of a
    merge but aren't the merge's top-left cell (openpyxl only stores the
    value there natively)."""
    values = {}
    for col in range(1, max_col + 1):
        v = ws.cell(row=row_num, column=col).value
        if v is not None:
            values[col] = v
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row_num <= mc.max_row:
            top_val = ws.cell(row=mc.min_row, column=mc.min_col).value
            for col in range(mc.min_col, mc.max_col + 1):
                if values.get(col) is None:
                    values[col] = top_val
    return values


def _build_subject_sheet_column_map(ws):
    """Maps every column to what it represents, based on the school's
    3-row header (exam section label, subject/field name, Marks/Grade
    sub-label). Distinguishes real per-subject columns (which have a
    literal 'Marks'/'Grade' in row 5) from the summary columns
    (Attendance, total marks, %, overall grade — which share row 5 with
    row 4 only because of a vertical merge, not a real sub-label)."""
    max_col = ws.max_column
    row4 = _fill_merged_row(ws, 4, max_col)
    row5_raw = {c: ws.cell(row=5, column=c).value for c in range(1, max_col + 1)}
    row3_raw = {c: ws.cell(row=3, column=c).value for c in range(1, max_col + 1)}

    section_labels = sorted([(c, v) for c, v in row3_raw.items() if v], key=lambda x: x[0])

    def section_for_col(col):
        sec = None
        for c, v in section_labels:
            if c <= col:
                sec = v
            else:
                break
        return sec

    column_map = []
    for col in range(1, max_col + 1):
        field = row4.get(col)
        subfield_raw = row5_raw.get(col)
        section_raw = section_for_col(col)
        section = None
        if section_raw:
            section = "FA1" if "1" in str(section_raw) else "FA2"

        if col == 1 and field and "sr" in str(field).lower():
            column_map.append({"col": col, "type": "sr_no"})
        elif col == 2 and field and "name" in str(field).lower():
            column_map.append({"col": col, "type": "name"})
        elif col == 3 and field and "mobile" in str(field).lower():
            column_map.append({"col": col, "type": "mobile"})
        elif col == 4 and field and "standard" in str(field).lower():
            column_map.append({"col": col, "type": "standard"})
        elif subfield_raw and str(subfield_raw).strip().lower() in ("marks", "grade"):
            column_map.append({
                "col": col, "type": "subject",
                "subject": str(field).strip(), "subfield": str(subfield_raw).strip().lower(),
                "section": section
            })
        elif field:
            field_lower = str(field).lower()
            if "attend" in field_lower:
                column_map.append({"col": col, "type": "attendance", "section": section})
            elif "marks" in field_lower:
                column_map.append({"col": col, "type": "total_marks", "section": section})
            elif str(field).strip() == "%":
                column_map.append({"col": col, "type": "percentage", "section": section})
            elif "grade" in field_lower:
                column_map.append({"col": col, "type": "overall_grade", "section": section})

    return column_map


def sync_subject_sheet(path):
    """
    Syncs the school's rich per-subject master sheet (.xlsx) — one row per
    student, with FA1/FA2 sections each broken into per-subject Marks+Grade
    columns (English, Kannada, Hindi, Maths, Science/EVS, Social), plus an
    overall Attendance/Total/%/Grade per section.

    Each subject becomes its own exam record (e.g. "FA1-MATHS", "FA1-ENG"),
    so parents can ask about a specific subject. The overall total per
    section is stored as its own record too (e.g. "FA1"), same as before.

    Safe to re-run — matches existing students by name+mobile, updates
    in place rather than duplicating.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    column_map = _build_subject_sheet_column_map(ws)

    col_lookup = {c["col"]: c for c in column_map}
    name_col = next((c["col"] for c in column_map if c["type"] == "name"), None)
    mobile_col = next((c["col"] for c in column_map if c["type"] == "mobile"), None)
    standard_col = next((c["col"] for c in column_map if c["type"] == "standard"), None)

    if not name_col or not mobile_col:
        print("Couldn't find 'Student Name' or 'Mobile No' columns. Aborting.")
        return

    subject_cols = [c for c in column_map if c["type"] == "subject"]
    sections_found = sorted(set(c["section"] for c in subject_cols if c["section"]))
    print(f"Detected sections: {sections_found}")
    subjects_found = sorted(set(c["subject"] for c in subject_cols))
    print(f"Detected subjects: {subjects_found}")

    # Find the real last data row (row 6 onward)
    last_row = 5
    for row in range(6, ws.max_row + 1):
        if ws.cell(row=row, column=name_col).value is not None:
            last_row = row
    total = last_row - 5
    print(f"Processing {total} students (using a single database connection for speed)...")

    def clean_str(val):
        """Handles Excel's habit of storing whole numbers as floats
        (e.g. 918147645010.0 instead of '918147645010')."""
        if val is None:
            return ""
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val).strip()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT roll_number FROM students WHERE roll_number LIKE 'MKS%'")
    existing_numbers = []
    for row in cur.fetchall():
        try:
            existing_numbers.append(int(row["roll_number"][3:]))
        except (ValueError, IndexError, TypeError):
            continue
    next_auto_num = (max(existing_numbers) + 1) if existing_numbers else 1

    created = 0
    updated = 0
    skipped = 0
    reference_rows = []

    for i, row_num in enumerate(range(6, last_row + 1), start=1):
        name = clean_str(ws.cell(row=row_num, column=name_col).value)
        mobile = clean_str(ws.cell(row=row_num, column=mobile_col).value)
        standard = clean_str(ws.cell(row=row_num, column=standard_col).value) if standard_col else ""

        if not name or not mobile:
            skipped += 1
            continue
        if not (mobile.isdigit() and len(mobile) >= 10):
            print(f"Skipping row with malformed mobile number for '{name}': '{mobile}'")
            skipped += 1
            continue

        grade_label = f"Standard {standard}" if standard else None

        cur.execute("""
            SELECT s.* FROM parents p JOIN students s ON p.student_id = s.id
            WHERE p.whatsapp_number = %s AND LOWER(s.student_name) = LOWER(%s)
        """, (mobile, name))
        existing = cur.fetchone()

        if existing:
            if grade_label is not None:
                cur.execute("UPDATE students SET grade = %s WHERE id = %s", (grade_label, existing["id"]))
            student_id = existing["id"]
            roll_number = existing["roll_number"]
            updated += 1
        else:
            roll_number = f"MKS{next_auto_num:04d}"
            next_auto_num += 1
            cur.execute(
                "INSERT INTO students (student_name, grade, roll_number) "
                "VALUES (%s, %s, %s) RETURNING id",
                (name, grade_label, roll_number)
            )
            student_id = cur.fetchone()["id"]
            cur.execute(
                "INSERT INTO parents (whatsapp_number, student_id) VALUES (%s, %s)",
                (mobile, student_id)
            )
            created += 1

        # Per-subject marks, per section (e.g. FA1-MATHS)
        for section in sections_found:
            section_subjects = {}
            for c in subject_cols:
                if c["section"] != section:
                    continue
                subj = c["subject"]
                val = ws.cell(row=row_num, column=c["col"]).value
                section_subjects.setdefault(subj, {})[c["subfield"]] = val

            for subj, vals in section_subjects.items():
                marks_val = vals.get("marks")
                grade_val = vals.get("grade")
                if marks_val is None:
                    continue  # no real data for this subject/section yet
                exam_name = f"{section}-{subj}".upper().replace(" ", "").replace("/", "")
                cur.execute("""
                    INSERT INTO marks (student_id, exam_name, marks, percentage, grade)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (student_id, exam_name)
                    DO UPDATE SET marks = EXCLUDED.marks, grade = EXCLUDED.grade,
                                   recorded_at = CURRENT_TIMESTAMP
                """, (student_id, exam_name, clean_str(marks_val), None, clean_str(grade_val) or None))

            # Overall section summary (e.g. "FA1") — attendance, total, %, grade
            section_cols = {c["type"]: c["col"] for c in column_map if c.get("section") == section and c["type"] != "subject"}
            attendance_val = ws.cell(row=row_num, column=section_cols["attendance"]).value if "attendance" in section_cols else None
            total_val = ws.cell(row=row_num, column=section_cols["total_marks"]).value if "total_marks" in section_cols else None
            pct_val = ws.cell(row=row_num, column=section_cols["percentage"]).value if "percentage" in section_cols else None
            overall_grade_val = ws.cell(row=row_num, column=section_cols["overall_grade"]).value if "overall_grade" in section_cols else None

            if attendance_val is not None:
                cur.execute("UPDATE students SET attendance = %s WHERE id = %s", (clean_str(attendance_val), student_id))

            if total_val is not None:
                cur.execute("""
                    INSERT INTO marks (student_id, exam_name, marks, percentage, grade)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (student_id, exam_name)
                    DO UPDATE SET marks = EXCLUDED.marks, percentage = EXCLUDED.percentage,
                                   grade = EXCLUDED.grade, recorded_at = CURRENT_TIMESTAMP
                """, (student_id, section, clean_str(total_val), clean_str(pct_val) or None, clean_str(overall_grade_val) or None))

        reference_rows.append((roll_number, name, mobile, grade_label or ""))

        if i % 50 == 0 or i == total:
            print(f"  ...{i}/{total} processed")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nSync complete: {created} new students added, {updated} existing students updated, {skipped} rows skipped.")

    ref_path = "assigned_ids_reference.csv"
    with open(ref_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["roll_number", "student_name", "parent_phone", "grade"])
        writer.writerows(reference_rows)
    print(f"Updated ID reference sheet: {ref_path}")


def import_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required_cols = {"parent_phone", "student_name", "grade", "attendance", "exam_schedule", "fee_status"}
        if not required_cols.issubset(reader.fieldnames or []):
            print(f"CSV is missing required columns. Needs: {sorted(required_cols)}")
            return
        all_rows = list(reader)

    has_roll_number = "roll_number" in (reader.fieldnames or []) and any(
        row.get("roll_number", "").strip() for row in all_rows
    )

    valid_rows = []
    skipped = 0
    for row in all_rows:
        if not row["student_name"].strip() or not row["parent_phone"].strip():
            print(f"Skipping row with missing name/phone: {row}")
            skipped += 1
            continue
        valid_rows.append(row)

    auto_ids = None
    if not has_roll_number:
        conn = get_connection()
        auto_ids = _next_auto_roll_numbers(conn, len(valid_rows))
        conn.close()
        print(f"No roll numbers found in CSV — auto-assigning {len(auto_ids)} school IDs (MKS0001 style).")

    added = 0
    reference_rows = []  # (roll_number, student_name, parent_phone, grade) — for the handoff sheet

    for i, row in enumerate(valid_rows):
        roll_number = row.get("roll_number", "").strip() if has_roll_number else auto_ids[i]
        add_student(
            row["parent_phone"], row["student_name"], row["grade"],
            row["attendance"], row["exam_schedule"], row["fee_status"],
            roll_number=roll_number
        )
        reference_rows.append((roll_number, row["student_name"], row["parent_phone"], row["grade"]))
        added += 1

    print(f"Import complete: {added} added, {skipped} skipped.")

    if not has_roll_number and reference_rows:
        ref_path = "assigned_ids_reference.csv"
        with open(ref_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["roll_number", "student_name", "parent_phone", "grade"])
            writer.writerows(reference_rows)
        print(f"\nSaved ID reference sheet: {ref_path}")
        print("Keep this file — it's what maps each student to their school ID.")
        print("Future mark uploads (FA2, SA1, etc.) need this 'roll_number' column filled in using this sheet.")


def update_marks_csv(path):
    """Bulk add/update exam marks for EXISTING students, matched by roll_number.
    Does not create new students — a roll_number with no matching student
    is reported and skipped, not silently dropped."""
    updated = 0
    not_found = []
    skipped = 0

    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required_cols = {"roll_number", "exam_name"}
        if not required_cols.issubset(reader.fieldnames or []):
            print(f"CSV is missing required columns. Needs at least: {sorted(required_cols)}")
            return

        for row in reader:
            roll_number = row.get("roll_number", "").strip()
            exam_name = row.get("exam_name", "").strip()

            if not roll_number or not exam_name:
                print(f"Skipping row with missing roll_number/exam_name: {row}")
                skipped += 1
                continue

            success = upsert_marks(
                roll_number=roll_number,
                exam_name=exam_name,
                marks=row.get("marks", "").strip() or None,
                percentage=row.get("percentage", "").strip() or None,
                grade=row.get("grade", "").strip() or None,
            )

            if success:
                updated += 1
            else:
                not_found.append(roll_number)

    print(f"Marks update complete: {updated} updated, {skipped} skipped (bad rows).")
    if not_found:
        print(f"No student found for {len(not_found)} roll number(s): {not_found}")


if __name__ == "__main__":
    init_db()

    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    command = sys.argv[1]

    if command == "add":
        add_student_interactive()
    elif command == "list":
        list_students()
    elif command == "sync-sheet":
        if len(sys.argv) < 3:
            print("Usage: python3 manage_students.py sync-sheet <path-to-csv>")
            sys.exit(1)
        sync_master_sheet(sys.argv[2])
    elif command == "sync-subjects":
        if len(sys.argv) < 3:
            print("Usage: python3 manage_students.py sync-subjects <path-to-xlsx>")
            sys.exit(1)
        sync_subject_sheet(sys.argv[2])
    elif command == "import":
        if len(sys.argv) < 3:
            print("Usage: python3 manage_students.py import <path-to-csv>")
            sys.exit(1)
        import_csv(sys.argv[2])
    elif command == "update-marks":
        if len(sys.argv) < 3:
            print("Usage: python3 manage_students.py update-marks <path-to-csv>")
            sys.exit(1)
        update_marks_csv(sys.argv[2])
    elif command == "remove":
        if len(sys.argv) < 3:
            print("Usage: python3 manage_students.py remove <student_id> [<student_id> ...]")
            sys.exit(1)
        for id_str in sys.argv[2:]:
            try:
                sid = int(id_str)
            except ValueError:
                print(f"Skipping invalid ID: '{id_str}'")
                continue
            name = delete_student(sid)
            if name:
                print(f"Removed student ID {sid} ({name})")
            else:
                print(f"No student found with ID {sid}")
    else:
        print(f"Unknown command: {command}")
        print(__doc__)